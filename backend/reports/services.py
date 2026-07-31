"""
Report generation services — produce PDF and Excel documents from live data.

PDFs are built with reportlab (platypus); spreadsheets with openpyxl. Every
function returns raw ``bytes`` so the views can stream them as downloads.
All data is read from the real PostgreSQL tables used across the app.
"""
import io
from datetime import datetime

from django.db import connection

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BRAND = colors.HexColor('#6C63FF')
BRAND_DARK = colors.HexColor('#4F46E5')
LIGHT = colors.HexColor('#EEF0FF')


# ──────────────────────────────────────────────────────────────────────────
# Data access
# ──────────────────────────────────────────────────────────────────────────
def _dictfetchall(cur):
    cols = [c[0] for c in cur.description]
    return [dict(zip(cols, r)) for r in cur.fetchall()]


def resolve_student(identifier):
    """Return a student dict for a student_id / user_id / email, or None."""
    with connection.cursor() as cur:
        cur.execute("""
            SELECT CAST(s.student_id AS TEXT) AS student_id, s.enrollment_no,
                   s.first_name, s.last_name, u.email,
                   d.name AS department_name, d.code AS department_code,
                   sem.number AS semester
            FROM students s
            LEFT JOIN users u ON CAST(u.id AS TEXT) = CAST(s.user_id AS TEXT)
            LEFT JOIN departments d ON CAST(d.department_id AS TEXT) = CAST(s.department_id AS TEXT)
            LEFT JOIN semesters sem ON CAST(sem.semester_id AS TEXT) = CAST(s.current_semester_id AS TEXT)
            WHERE CAST(s.student_id AS TEXT) = %s
               OR CAST(s.user_id AS TEXT) = %s
               OR LOWER(u.email) = LOWER(%s)
            LIMIT 1
        """, [str(identifier), str(identifier), str(identifier)])
        rows = _dictfetchall(cur)
    return rows[0] if rows else None


def _student_marks(student_id):
    with connection.cursor() as cur:
        cur.execute("""
            SELECT sub.code, sub.name, sub.credits, sem.number AS semester,
                   m.marks, m.grade, m.grade_points
            FROM marks m
            LEFT JOIN subjects sub ON CAST(sub.subject_id AS TEXT) = CAST(m.subject_id AS TEXT)
            LEFT JOIN semesters sem ON CAST(sem.semester_id AS TEXT) = CAST(m.semester_id AS TEXT)
            WHERE CAST(m.student_id AS TEXT) = %s
            ORDER BY sem.number, sub.code
        """, [str(student_id)])
        return _dictfetchall(cur)


def _student_attendance(student_id):
    with connection.cursor() as cur:
        cur.execute("""
            SELECT sub.code, sub.name, a.total_lectures, a.present_count,
                   a.percentage, a.is_shortage
            FROM attendance_summary a
            LEFT JOIN subjects sub ON CAST(sub.subject_id AS TEXT) = CAST(a.subject_id AS TEXT)
            WHERE CAST(a.student_id AS TEXT) = %s
            ORDER BY sub.code
        """, [str(student_id)])
        return _dictfetchall(cur)


def _student_fees(student_id):
    with connection.cursor() as cur:
        cur.execute("""
            SELECT fs.component_name, fs.amount, fp.amount_paid,
                   fp.status, fp.payment_date, fp.transaction_ref
            FROM fee_payments fp
            LEFT JOIN fee_structures fs ON CAST(fs.fee_id AS TEXT) = CAST(fp.fee_structure_id AS TEXT)
            WHERE CAST(fp.student_id AS TEXT) = %s
            ORDER BY fp.payment_date DESC NULLS LAST
        """, [str(student_id)])
        return _dictfetchall(cur)


# ──────────────────────────────────────────────────────────────────────────
# Shared PDF helpers
# ──────────────────────────────────────────────────────────────────────────
def _styles():
    ss = getSampleStyleSheet()
    ss.add(ParagraphStyle('Brand', parent=ss['Title'], textColor=BRAND_DARK, fontSize=20, spaceAfter=2))
    ss.add(ParagraphStyle('Sub', parent=ss['Normal'], textColor=colors.grey, fontSize=9, spaceAfter=10))
    ss.add(ParagraphStyle('H', parent=ss['Heading2'], textColor=BRAND_DARK, fontSize=12, spaceBefore=10, spaceAfter=6))
    return ss


def _header(story, ss, title, student):
    story.append(Paragraph('College360 Pro', ss['Brand']))
    story.append(Paragraph(f'{title} &nbsp;·&nbsp; generated {datetime.now():%d %b %Y, %H:%M}', ss['Sub']))
    if student:
        info = [
            ['Name', f"{student.get('first_name','')} {student.get('last_name','')}".strip(), 'Enrollment', student.get('enrollment_no') or '—'],
            ['Department', student.get('department_name') or '—', 'Email', student.get('email') or '—'],
        ]
        t = Table(info, colWidths=[26*mm, 60*mm, 26*mm, 55*mm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (0,-1), LIGHT), ('BACKGROUND', (2,0), (2,-1), LIGHT),
            ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'), ('FONTNAME', (2,0), (2,-1), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 9), ('BOX', (0,0), (-1,-1), 0.5, colors.lightgrey),
            ('INNERGRID', (0,0), (-1,-1), 0.5, colors.lightgrey), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 5), ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ]))
        story.append(t)
        story.append(Spacer(1, 8))


def _data_table(rows, header, col_widths, aligns=None):
    data = [header] + rows
    t = Table(data, colWidths=col_widths, repeatRows=1)
    style = [
        ('BACKGROUND', (0,0), (-1,0), BRAND), ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'), ('FONTSIZE', (0,0), (-1,-1), 8.5),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, LIGHT]),
        ('BOX', (0,0), (-1,-1), 0.5, colors.lightgrey), ('INNERGRID', (0,0), (-1,-1), 0.4, colors.lightgrey),
        ('TOPPADDING', (0,0), (-1,-1), 4), ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]
    for col, a in (aligns or {}).items():
        style.append(('ALIGN', (col,0), (col,-1), a))
    t.setStyle(TableStyle(style))
    return t


# ──────────────────────────────────────────────────────────────────────────
# PDF: marksheet
# ──────────────────────────────────────────────────────────────────────────
def generate_marksheet_pdf(identifier):
    student = resolve_student(identifier)
    if not student:
        return None
    marks = _student_marks(student['student_id'])

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=18*mm, bottomMargin=18*mm,
                            leftMargin=16*mm, rightMargin=16*mm, title='Marksheet')
    ss = _styles()
    story = []
    _header(story, ss, 'Academic Marksheet', student)

    if not marks:
        story.append(Paragraph('No marks recorded for this student yet.', ss['Normal']))
        doc.build(story)
        return buf.getvalue()

    # group by semester, compute SGPA
    by_sem = {}
    for m in marks:
        by_sem.setdefault(m.get('semester') or 0, []).append(m)

    total_cr = total_wcr = 0.0
    for sem in sorted(by_sem):
        rows, sem_cr, sem_wcr = [], 0.0, 0.0
        for m in by_sem[sem]:
            cr = float(m.get('credits') or 0)
            gp = float(m.get('grade_points') or 0)
            sem_cr += cr; sem_wcr += cr * gp
            rows.append([m.get('code') or '—', (m.get('name') or '')[:38],
                         f"{cr:g}", str(m.get('marks') or '—'), m.get('grade') or '—', f"{gp:g}"])
        sgpa = round(sem_wcr / sem_cr, 2) if sem_cr else 0.0
        total_cr += sem_cr; total_wcr += sem_wcr
        story.append(Paragraph(f'Semester {sem}  —  SGPA {sgpa}', ss['H']))
        story.append(_data_table(rows, ['Code', 'Subject', 'Cr', 'Marks', 'Grade', 'GP'],
                                 [22*mm, 66*mm, 14*mm, 20*mm, 20*mm, 16*mm],
                                 aligns={2:'CENTER', 3:'CENTER', 4:'CENTER', 5:'CENTER'}))
        story.append(Spacer(1, 4))

    cgpa = round(total_wcr / total_cr, 2) if total_cr else 0.0
    story.append(Spacer(1, 8))
    summary = _data_table([[f'{total_cr:g}', str(len(marks)), str(cgpa)]],
                          ['Total Credits', 'Subjects', 'CGPA'],
                          [45*mm, 45*mm, 45*mm], aligns={0:'CENTER',1:'CENTER',2:'CENTER'})
    story.append(summary)
    doc.build(story)
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────────────────
# PDF: fee statement
# ──────────────────────────────────────────────────────────────────────────
def generate_fee_statement_pdf(identifier):
    student = resolve_student(identifier)
    if not student:
        return None
    fees = _student_fees(student['student_id'])

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=18*mm, bottomMargin=18*mm,
                            leftMargin=16*mm, rightMargin=16*mm, title='Fee Statement')
    ss = _styles()
    story = []
    _header(story, ss, 'Fee Statement', student)

    rows, total_due, total_paid = [], 0.0, 0.0
    for f in fees:
        amt = float(f.get('amount') or 0)
        paid = float(f.get('amount_paid') or 0)
        total_due += amt; total_paid += paid
        pdate = f.get('payment_date')
        rows.append([(f.get('component_name') or 'Fee')[:34], f"{amt:,.0f}", f"{paid:,.0f}",
                     (f.get('status') or '—').title(),
                     pdate.strftime('%d-%m-%Y') if hasattr(pdate, 'strftime') else (str(pdate) if pdate else '—')])
    if not rows:
        story.append(Paragraph('No fee records for this student.', ss['Normal']))
    else:
        story.append(_data_table(rows, ['Component', 'Amount', 'Paid', 'Status', 'Date'],
                                 [58*mm, 26*mm, 26*mm, 26*mm, 26*mm],
                                 aligns={1:'RIGHT', 2:'RIGHT', 3:'CENTER', 4:'CENTER'}))
        story.append(Spacer(1, 10))
        bal = total_due - total_paid
        story.append(_data_table([[f"{total_due:,.0f}", f"{total_paid:,.0f}", f"{bal:,.0f}"]],
                                 ['Total Billed', 'Total Paid', 'Balance Due'],
                                 [45*mm, 45*mm, 45*mm], aligns={0:'RIGHT',1:'RIGHT',2:'RIGHT'}))
    doc.build(story)
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────────────────
# Excel helpers
# ──────────────────────────────────────────────────────────────────────────
_HEAD_FILL = PatternFill('solid', fgColor='6C63FF')
_HEAD_FONT = Font(bold=True, color='FFFFFF')
_THIN = Side(style='thin', color='D0D0E0')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _style_sheet(ws, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = _HEAD_FILL; c.font = _HEAD_FONT
        c.alignment = Alignment(horizontal='center', vertical='center')
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ncols):
        for c in row:
            c.border = _BORDER
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 3, 12), 45)


def generate_attendance_excel(identifier):
    student = resolve_student(identifier)
    if not student:
        return None
    rows = _student_attendance(student['student_id'])
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Attendance'
    ws.append(['Subject Code', 'Subject', 'Lectures', 'Present', 'Percentage', 'Shortage?'])
    for r in rows:
        ws.append([r.get('code') or '—', r.get('name') or '—', r.get('total_lectures') or 0,
                   r.get('present_count') or 0, float(r.get('percentage') or 0),
                   'YES' if r.get('is_shortage') else 'No'])
    if rows:
        avg = round(sum(float(r.get('percentage') or 0) for r in rows) / len(rows), 1)
        ws.append([]); ws.append(['', 'OVERALL', '', '', avg, ''])
    _style_sheet(ws, 6)
    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()


def generate_department_summary_excel():
    """Institute-wide department summary: headcount, avg CPI, avg attendance."""
    with connection.cursor() as cur:
        cur.execute("""
            SELECT d.name AS department, d.code,
                   COUNT(DISTINCT s.student_id) AS students,
                   ROUND(AVG(msub.cpi)::numeric, 2) AS avg_cpi,
                   ROUND(AVG(asub.att)::numeric, 1) AS avg_attendance
            FROM departments d
            LEFT JOIN students s ON CAST(s.department_id AS TEXT) = CAST(d.department_id AS TEXT)
            LEFT JOIN (
                SELECT student_id, AVG(grade_points) AS cpi FROM marks GROUP BY student_id
            ) msub ON CAST(msub.student_id AS TEXT) = CAST(s.student_id AS TEXT)
            LEFT JOIN (
                SELECT student_id, AVG(percentage) AS att FROM attendance_summary GROUP BY student_id
            ) asub ON CAST(asub.student_id AS TEXT) = CAST(s.student_id AS TEXT)
            GROUP BY d.name, d.code
            ORDER BY students DESC
        """)
        rows = _dictfetchall(cur)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Departments'
    ws.append(['Department', 'Code', 'Students', 'Avg CPI', 'Avg Attendance %'])
    for r in rows:
        ws.append([r['department'], r['code'], r['students'] or 0,
                   float(r['avg_cpi'] or 0), float(r['avg_attendance'] or 0)])
    _style_sheet(ws, 5)
    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()
